import streamlit as st
from datetime import date, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import re
import time
import os
from openpyxl import load_workbook
import json

# ----------------------------
# 0️⃣ 안내 박스
# ----------------------------
st.title("최저가 자동수집 & 엑셀 업데이트")

st.info("""
🔹 **주의사항 및 안내**

🔹 크롤링(수집 주소)의 위치가 변동되면 파일이 깨질 수 있으므로 백업용 파일은 필수로 보관해야 합니다. 

1. **하이디럭스 객실은 자동 수집이 어렵습니다.**  
   → 수집 후 가격은 반드시 **엑셀에서 직접 입력**해주세요.

2. **엑셀 양식 주의**
   - 시트명은 `{년}.{월}` 형식이어야 합니다. (예: 26.03)
   - 첫 번째 컬럼은 **일(day)**이어야 합니다.
   - 호텔명은 기존 컬럼 순서를 유지해야 합니다.
   - 엑셀 파일 위치: `\\객실프론트1\공유문서\객실팀 공유폴더\야간조폴더\온라인 경쟁 호텔.xlsx`

3. **사용 방법**
   1. 날짜 선택
   2. "최저가 수집 & 엑셀 업데이트" 버튼 클릭
   3. 자동 수집 완료 후, 하이디럭스 가격을 엑셀에 수동 입력
   4. 엑셀 저장 확인
""")

# ----------------------------
# 1️⃣ 날짜 선택
# ----------------------------
selected_date = st.date_input("날짜 선택", value=date.today())
checkin = selected_date.strftime("%Y-%m-%d")
checkout = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

# ----------------------------
# 2️⃣ 엑셀 파일 및 시트 선택
# ----------------------------
excel_path = r"\\200.1.74.202\공유문서\객실팀 공유폴더\야간조폴더\온라인 경쟁 호텔.xlsx"
sheet_name = f"{str(selected_date.year)[-2:]}.{selected_date.month:02d}"

if not os.path.exists(excel_path):
    st.error(f"엑셀 파일이 없습니다: {excel_path}")
else:
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            st.error(f"{sheet_name} 시트가 없습니다.")
        else:
            ws = wb[sheet_name]
            st.write(f"{sheet_name} 시트 로드 완료")
    except Exception as e:
        st.error(f"엑셀 로드 실패: {e}")

    # ----------------------------
    # 3️⃣ 호텔 ID 정의
    # ----------------------------
    hotel_ids = {
        "파라다이스": "11576700",
        "파크하얏트": "31696883",
        "롯데": "11577921",
        "웨스틴조선": "11565526",
        "아난티 앳 부산 코브": "650968109",
        "시그니엘부산": "1857519376",
        "그랜드 조선": "1016404332",
        "아바니 센트럴(문현동)": "1431665023",
        "신라스테이(해운대)": "791821570",
        "신라스테이(서부산)": "1568839412",
        "농심(디럭스)": "11555977",
        "농심(하이디럭스)": "11555977"
    }

    # ----------------------------
    # 4️⃣ URL 생성 함수
    # ----------------------------
    def make_url(hotel_id, checkin, checkout, type_key=None):
        base = f"https://hotels.naver.com/accommodation/search/detail/domestic/{hotel_id}/rates?dAdultCnt=2&dCheckIn={checkin}&dCheckOut={checkout}"
        if type_key == "하이디럭스":
            base += "&rateTab=&keyword=detail&lodgingType=domestic&slug={}".format(hotel_id)
        return base

    # ----------------------------
    # 5️⃣ 가격 수집 버튼
    # ----------------------------
    if st.button("최저가 수집 & 엑셀 업데이트"):
        st.write("Selenium 실행 중... 잠시만 기다려주세요.")

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

        # ----------------------------
        # DATE 기준 행 찾기
        # ----------------------------
        date_found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == selected_date.day:
                target_row = row
                date_found = True
                break

        if not date_found:
            ws.append([selected_date.day] + [None]*(ws.max_column-1))
            ws_row_idx = ws.max_row
        else:
            ws_row_idx = target_row[0].row

        # ----------------------------
        # 진행 상태 표시
        # ----------------------------
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_hotels = len(hotel_ids)

        hotel_list = list(hotel_ids.keys())

        # ----------------------------
        # 호텔별 가격 수집
        # ----------------------------
        for idx, hotel in enumerate(hotel_list):

            col_idx = idx + 3
            type_key = None

            if "농심(하이디럭스)" in hotel:
                type_key = "하이디럭스"

            url = make_url(hotel_ids[hotel], checkin, checkout, type_key)
            driver.get(url)

            status_text.text(f"{hotel} 가격 수집 중... ({idx+1}/{total_hotels})")

            try:
                if type_key == "하이디럭스":

                    ws.cell(row=ws_row_idx, column=col_idx, value="직접입력")
                    status_text.text(f"{hotel}: 하이디럭스 직접 입력 필요")

                else:

                    price_elements = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located(
                            (By.CSS_SELECTOR, "div.common_PriceInfo__vuU30 em.common_price__iYQ6j")
                        )
                    )

                    prices = []

                    for el in price_elements:
                        try:
                            num = int(re.sub("[^0-9]", "", el.text))
                            prices.append(num)
                        except:
                            continue

                    if prices:
                        ws.cell(row=ws_row_idx, column=col_idx, value=min(prices))
                        status_text.text(f"{hotel}: {min(prices)}원 입력 완료")

                    else:
                        ws.cell(row=ws_row_idx, column=col_idx, value="마감")
                        status_text.text(f"{hotel}: 마감")

            except Exception as e:

                ws.cell(row=ws_row_idx, column=col_idx, value="마감")
                status_text.text(f"{hotel}: 마감 (오류 발생)")

            progress_bar.progress((idx+1) / total_hotels)
            time.sleep(1)

        driver.quit()

        # ----------------------------
        # 6️⃣ 엑셀 저장
        # ----------------------------
        wb.save(excel_path)
        st.success("엑셀 업데이트 완료!")
